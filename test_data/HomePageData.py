import openpyxl


class HomePageData:

    # test_homepage_data = [{"fullname": "Abraham Ora", "email": "abraham.morial.ora@gmail.com", "gender": "Male"},
    #                         {"fullname": "Clarisse Ora", "email": "clarisse.licuanan.ora@gmail.com", "gender": "Female"}
    #                         ]

    @staticmethod
    def getTestData(test_case_name):
        dict = {}
        book = openpyxl.load_workbook("/Users/hnstabe/Downloads/PythonDemo.xlsx")
        sheet = book.active

        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [dict]
